from openpyxl import load_workbook
from util import _get_cell_value, get_boundary_values
import files_path


class CommonFunctions:
    def __init__(self):
        self.file_path = files_path.agent_report
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)


class WorkSheet(CommonFunctions):
    def table_range(self, table):
        """Возвращает диапазон выбранной таблицы

        :param table: 'tickets' или 'receipts'
        :return: string
        """
        # задаем начальное и конечное значение для поиска диапазона
        start_value_row = 'Наименование филиала'
        end_value_row = 'ИТОГО:'
        end_value_column = 'Подлежит перечислению Принципалу за отчетный период'
        table_range = [start_value_row, end_value_column, end_value_row]

        # Функция iter_rows() возвращает все строки из файла
        rows = self.sheet.iter_rows()
        # Переменная `cells` содержит генератор, который является последовательностью всех ячеек в листе таблицы
        cells = (cell for row in rows for cell in row)
        # Поиск стартовой точки, последней колонки и строки
        ranges_in_array = (cell for cell in cells if cell.value in table_range)

        # Диапазон первой таблицы
        start_position = next(ranges_in_array)
        last_column, last_row = next(ranges_in_array), next(ranges_in_array)
        # Диапазон второй таблицы
        start_position_st = next(ranges_in_array)
        last_column_st, last_row_st = next(ranges_in_array), next(ranges_in_array)

        coordinate = {
            'start_row': start_position.row + 4 if table == 'tickets' else start_position_st.row + 4,
            'start_column': start_position.column if table == 'tickets' else start_position_st.column,
            'last_column': last_column.column if table == 'tickets' else last_column_st.column + 4,
            'last_row': last_row.row if table == 'tickets' else last_row_st.row
        }

        start_coordinate = self.sheet.cell(row=coordinate["start_row"], column=coordinate["start_column"]).coordinate
        end_coordinate = self.sheet.cell(row=coordinate["last_row"], column=coordinate["last_column"]).coordinate
        table_range = start_coordinate + ':' + end_coordinate

        return table_range


# Объекты класса WorkSheet
workSheet = WorkSheet()
tickets_table_range = workSheet.table_range('tickets')
receipts_table_range = workSheet.table_range('receipts')


class ReportAgentData(CommonFunctions):
    # TODO: отрефакторить наподобии
    """
    def get_values_from_table(self, column, table_range):
        result_row = int(get_boundary_values(table_range, 'max_row'))
        values = {
            'sold_number': self._get_cell_value('C', result_row),
            'sold_amount': self._get_cell_value('E', result_row),
            'paid_number': self._get_cell_value('H', result_row),
            'paid_amount': self._get_cell_value('J', result_row),
            'reward': round(Decimal(self._get_cell_value('N', result_row)), 2),
            'transfer': round(Decimal(self._get_cell_value('O', result_row)), 2)
        }
        return values.get(column)

    def total_values_lottery_tickets(self, column):
        return self.get_values_from_table(column, tickets_table_range)

    def total_values_lottery_receipts(self, column):
        return self.get_values_from_table(column, receipts_table_range)
    """

    # Беру данные из "ИТОГО" в таблице "Реализация лотерейных билетов"
    def total_values_lottery_tickets(self, column):
        result_row = get_boundary_values(tickets_table_range, 'max_row')
        values = {
            # Продажи
            'sold_number': self._get_cell_value('F', result_row),
            'sold_amount': self._get_cell_value('G', result_row),
            # Выплаты
            'paid_number': self._get_cell_value('K', result_row),
            'paid_amount': self._get_cell_value('L', result_row),
            # Вознаграждения
            'reward': round(float(self._get_cell_value('N', result_row)), 2),
            # Перечисление принципалу
            'transfer': round(float(self._get_cell_value('O', result_row)), 2)
        }
        return values[column]

    # Беру данные из "ИТОГО" в таблице "Реализация лотерейных квитанций"
    def total_values_lottery_receipts(self, column):
        result_row = get_boundary_values(receipts_table_range, 'max_row')
        values = {
            # Продажи
            'sold_number': self._get_cell_value('C', result_row),
            'sold_amount': self._get_cell_value('E', result_row),
            # Выплаты
            'paid_number': self._get_cell_value('H', result_row),
            'paid_amount': self._get_cell_value('J', result_row),
            # Вознаграждения
            'reward': round(float(self._get_cell_value('N', result_row)), 2),
            # Перечисление принципалу
            'transfer': round(float(self._get_cell_value('O', result_row)), 2)
        }
        return values[column]

    def get_values_in_column(self, column, add_to_row):
        result_row = int(get_boundary_values(receipts_table_range, 'max_row')) + add_to_row
        reward = self._get_cell_value(column, result_row)
        return reward

    # Общая сумма двух таблиц по вознаграждению
    def get_reward_of_two_tables(self) -> float:
        summary = float(self.get_values_in_column('K', 5))
        return summary

    # Общая сумма двух таблиц по перечислению средств
    def get_transfer_of_two_tables(self) -> float:
        summary = float(self.get_values_in_column('K', 6))
        return summary

    # Общая сумма двух таблиц по продажам
    def get_sales_of_two_tables(self):
        return self.get_values_in_column('O', 3)

    # Общая сумма двух таблиц по выплатам
    def get_payment_of_two_tables(self):
        return self.get_values_in_column('O', 4)


class AgentAsserts(CommonFunctions):
    # Функция для подсчета итоговых данных
    # в таблицах "Реализация лотерейных билетов" и "Реализация лотерейных квитанций"
    def counting_values_in_column(self, table, column, data_type='int'):
        total = 0
        if table == 'realization_tickets':
            diapason = tickets_table_range
        elif table == 'realization_receipts':
            diapason = receipts_table_range
        else:
            raise ValueError('Неверный тип таблицы')

        min_row = int(get_boundary_values(diapason, 'min_row'))
        max_row = int(get_boundary_values(diapason, 'max_row'))

        for row in range(min_row, max_row):
            cell = self.sheet.cell(row=row, column=column).value
            if data_type == 'float':
                total += float(cell)
            else:
                total += int(cell)

        if data_type == 'float':
            return round(total, 2)
        else:
            return total

    # "Реализация лотерейных билетов"
    def sold_number_tickets(self):
        return self.counting_values_in_column('realization_tickets', 6)

    def sold_amount_tickets(self):
        return self.counting_values_in_column('realization_tickets', 7)

    def paid_number_tickets(self):
        return self.counting_values_in_column('realization_tickets', 11)

    def paid_amount_tickets(self):
        return self.counting_values_in_column('realization_tickets', 12)

    def reward_tickets(self):
        return self.counting_values_in_column('realization_tickets', 14, 'float')

    def transfer_tickets(self):
        return self.counting_values_in_column('realization_tickets', 15, 'float')

    # "Реализация лотерейных квитанций"
    def sold_number_receipts(self):
        return self.counting_values_in_column('realization_receipts', 3)

    def sold_amount_receipts(self):
        return self.counting_values_in_column('realization_receipts', 5)

    def paid_number_receipts(self):
        return self.counting_values_in_column('realization_receipts', 8)

    def paid_amount_receipts(self):
        return self.counting_values_in_column('realization_receipts', 10)

    def reward_receipts(self):
        return self.counting_values_in_column('realization_receipts', 14, 'float')

    def transfer_receipts(self):
        return self.counting_values_in_column('realization_receipts', 15, 'float')

    # Общее вознаграждение и перечисление
    def total_rewards(self) -> float:
        tickets = self.reward_tickets()
        receipts = self.reward_receipts()
        return tickets + receipts

    def total_transfer(self) -> float:
        tickets = self.transfer_tickets()
        receipts = self.transfer_receipts()
        return tickets + receipts

    # Проверка расчетов в каждой строке
    def check_row(self, table):
        if table == 'realization_tickets':
            diapason = tickets_table_range
        elif table == 'realization_receipts':
            diapason = receipts_table_range
        else:
            raise ValueError('Неверный тип таблицы')

        min_row = int(get_boundary_values(diapason, 'min_row'))
        max_row = int(get_boundary_values(diapason, 'max_row'))

        for row in range(min_row, max_row):
            columns = {
                'sold_amount_column': 7 if table == 'realization_tickets' else 5,
                'paid_amount_column': 11 if table == 'realization_tickets' else 8,
                'percent_column': 12 if table == 'realization_tickets' else 10,
                'reward_column': 14 if table == 'realization_tickets' else 14,
                'transfer_column': 15 if table == 'realization_tickets' else 15,
            }

            sold_amount = self.sheet.cell(row=row, column=columns['sold_amount_column']).value
            paid_amount = self.sheet.cell(row=row, column=columns['paid_amount_column']).value
            percent = self.sheet.cell(row=row, column=columns['percent_column']).value
            reward = float(self.sheet.cell(row=row, column=columns['reward_column']).value)
            transfer = float(self.sheet.cell(row=row, column=columns['transfer_column']).value)

            try:
                count_reward = round(sold_amount * (percent / 100), 1)
                assert count_reward == round(reward, 1)

                count_transfer = sold_amount - paid_amount - round(reward, 1)
                assert count_transfer == transfer

                return "ДА"
            except:
                return 'НЕТ!\nСтрока ' + str(row)
