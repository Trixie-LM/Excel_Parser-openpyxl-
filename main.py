from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from paid_payments_registry import PreCondition, ReportPaidPaymentRegistry
import discrepancies_in_reports
import files_path
from openpyxl import Workbook
from util import Timer
import common_core


# Вычисление время вызова приложения
timer = Timer()

# Переменные классов
#TODO: Удалить report_p_p_registry, импортировать из CC
report_p_p_registry = ReportPaidPaymentRegistry()
PreCondition = PreCondition()


class CreatingFinalReport:
    def __init__(self):
        self.file_path = files_path.total_report
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.title = "Итоговый отчет"

    def preconditions(self):
        # Создание короткого отчета
        PreCondition.copying_table()
        PreCondition.delete_rows()

        # Копирование таблицы "реестр выплаченных выигрышей" в основной файл
        for row in report_p_p_registry.paid_lottery_names_list():
            reversed_row = reversed(row)
            for cell in reversed_row:
                self.sheet.cell(row=cell.row + 3, column=cell.column + 20).value = cell.value

    # Редактирование клеток в файле
    def tables_editing(self):
        print('Соединяем ячейки и устанавливаем стиль...')
        for cell in common_core.editing_cells():
            double = Side(border_style="double", color="FF000000")
            startCell = cell.split(':')[0]

            self.sheet[startCell].font = Font(size=10)
            self.sheet[startCell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet[startCell].border = Border(top=double, bottom=double, left=double, right=double)
            self.sheet.merge_cells(cell)

    def import_data(self):
        print('Импортируем данные в таблицу...')
        # Импорт данных в таблицу
        for columnNumber, text in common_core.input_data():
            self.sheet[columnNumber].value = text

        timer.tick('Импорт данных занял')

    # TODO: найти решение лучше
    def get_column_wider(self):
        self.sheet.column_dimensions["F"].auto_size = True
        self.sheet.column_dimensions["P"].auto_size = True
        self.sheet.column_dimensions["U"].width = 12
        self.sheet.column_dimensions["V"].width = 12

    def check_and_paint(self):
        print('Проверяем значения между собой и красим ячейки...')
        # Сверка данных и заливка фона ячейки
        for cellNumber, numberInReport, verifiable in common_core.check_and_painting():
            if str(numberInReport) == str(verifiable):
                self.sheet[cellNumber].fill = PatternFill('solid', fgColor="00FF00")
            else:
                self.sheet[cellNumber].fill = PatternFill('solid', fgColor="FF0000")

        timer.tick('Проверка значений заняла')
        self.book.save(self.file_path)

    @staticmethod
    def postconditions():
        discrepancies_in_reports.postconditions()

    def calling_all_methods(self):
        self.preconditions()
        self.tables_editing()
        self.import_data()
        self.get_column_wider()
        self.check_and_paint()
        self.postconditions()
