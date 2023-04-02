from openpyxl import load_workbook
import files_path

document = files_path.report_of_checking

book = load_workbook(filename=document, data_only=True)
sheet = book.active


def amount_counting(char):
    summary = 0

    for i in range(2, sheet.max_row + 1):
        cellValue = sheet[char + str(i)].value
        summary += cellValue

    return summary


# Продажа
def number_of_circulation_sales():
    return amount_counting('E')

def amount_of_circulation_sales():
    return amount_counting('F')

def number_of_digital_sales():
    return amount_counting('G')

def amount_of_digital_sales():
    return amount_counting('H')

def number_of_instant_sales():
    return amount_counting('I')

def amount_of_instant_sales():
    return amount_counting('J')

def total_number_of_sales():
    return amount_counting('K')

def total_amount_of_sales():
    return amount_counting('L')

# Выплата
def number_of_circulation_payments():
    return amount_counting('M')

def amount_of_circulation_payments():
    return amount_counting('N')

def number_of_digital_payments():
    return amount_counting('O')

def amount_of_digital_payments():
    return amount_counting('P')

def number_of_instant_payments():
    return amount_counting('Q')

def amount_of_instant_payments():
    return amount_counting('R')

def total_number_of_payments():
    return amount_counting('S')

def total_amount_of_payments():
    return amount_counting('T')
