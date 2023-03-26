from openpyxl import load_workbook
from numpy import unique
from util import _get_cell_value


class SalesDataBase:
    def __init__(self):
        self.file_path = 'C:/Users/Trixie_LM/Desktop/1C/Проданные билеты.xlsx'
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)

class ReportSalesData(SalesDataBase):

    def total_quantity_tickets_in_report(self):
        return self._get_cell_value('M', self.sheet.max_row)

    def tickets_price_in_report(self):
        return self._get_cell_value('N', self.sheet.max_row)

    def check_cells_in_row(self):
        passedAssertions = 0

        for i in range(2, self.sheet.max_row):
            setCell = self._get_cell_value('L', i)
            drawCell = self._get_cell_value('I', i)
            ticketType = self._get_cell_value('K', i)
            lotteryType = self._get_cell_value('F', i)
            seventhDigit = self._get_cell_value('J', i)[6:7]

            try:
                if lotteryType == "Бинго лотерея" and seventhDigit == "3" and len(
                        drawCell) == 6 and setCell is not None:
                    assert ticketType == "Открытка"
                    passedAssertions += 1

                elif lotteryType == "Бинго лотерея" and seventhDigit == "3" and len(drawCell) == 6 and setCell is None:
                    assert ticketType == "Бумажный"
                    passedAssertions += 1

                elif lotteryType == "Моментальная лотерея" and setCell is None and drawCell is None:
                    assert ticketType == "Бумажный"
                    passedAssertions += 1

                elif seventhDigit == "1" and len(drawCell) == 6 and setCell is None:
                    assert ticketType == "Электронный"
                    passedAssertions += 1

                elif seventhDigit == "2" and len(drawCell) == 6 and setCell is None:
                    assert ticketType == "Купон"
                    passedAssertions += 1

                else:
                    print('Ошибка! Отчет продаж. Строка - ' + str(i))


            except AssertionError:
                print('Ошибка! Отчет продаж. Строка - ' + str(i))


        return passedAssertions


class SalesAsserts(SalesDataBase):

    def total_quantity(self):
        tickets_amount = 0
        for i in range(2, self.sheet.max_row):
            tickets_amount += self._get_cell_value('M', i)
        return tickets_amount

    def tickets_price(self):
        tickets_price = 0
        for i in range(2, self.sheet.max_row):
            tickets_price += self._get_cell_value('N', i)
        return tickets_price

    def counting_tickets(self, type_of_tickets, type_of_lottery=True):
        ticketsQuantity = 0
        ticketsPrice = 0
        bingoLottery = ["102021", "102041", "102051", "102091"]

        for i in range(2, self.sheet.max_row):
            productCode = self._get_cell_value('J', i)[0:6]
            ticketType = self._get_cell_value('K', i)
            ticketPrice = self._get_cell_value('N', i)

            # Подсчет электронных билетов и купонов
            if type_of_tickets == 'Электронный':
                if ticketType in ['Электронный', 'Купон']:
                    ticketsQuantity += 1
                    ticketsPrice += ticketPrice

            # Подсчет бумажных моменталок
            elif type_of_tickets == 'Бумажный' and type_of_lottery == 'Моментальная':
                if ticketType == type_of_tickets and productCode not in bingoLottery:
                    ticketsQuantity += 1
                    ticketsPrice += ticketPrice

            # Подсчет тиражных билетов
            elif type_of_tickets == 'Бумажный' and type_of_lottery == 'Тиражная':
                if ticketType in ['Бумажный', 'Открытка'] and productCode in bingoLottery:
                    ticketsQuantity += 1
                    ticketsPrice += ticketPrice

            # Подсчет бумажных бинго
            elif type_of_tickets == 'Бумажный':
                if ticketType == type_of_tickets:
                    ticketsQuantity += 1
                    ticketsPrice += ticketPrice

            # Подсчет отстальных (т.е. Открыток)
            else:
                if ticketType == type_of_tickets:
                    ticketsQuantity += 1
                    ticketsPrice += ticketPrice

        return ticketsQuantity, ticketsPrice

    def all_ticket_types(self):
        ticketsQuantity = self.counting_tickets('Электронный')[0] + self.counting_tickets('Бумажный')[0] + self.counting_tickets('Открытка')[0]
        ticketsPrice = self.counting_tickets('Электронный')[1] + self.counting_tickets('Бумажный')[1] + self.counting_tickets('Открытка')[1]

        return ticketsQuantity, ticketsPrice

    def unique_ticket_numbers(self):
        ticketNumbersQuantity = 0
        ticketNumbers = []
        for i in range(2, self.sheet.max_row):
            ticketNumber = self._get_cell_value('J', i)
            if ticketNumber is not None:
                ticketNumbers.append(ticketNumber)
                ticketNumbersQuantity += 1

        if ticketNumbersQuantity == len(unique(ticketNumbers)):
            return "Все билеты в отчете уникальные"
        else:
            return "В отчете есть дубликат билета"

    def tickets_of_sets_in_report(self):
        setAmount = 0
        allSets = []
        for i in range(2, self.sheet.max_row):

            ticketSetNumber = self._get_cell_value('L', i)
            if ticketSetNumber is not None:
                allSets.append(ticketSetNumber)
                setAmount += 1

        if setAmount % 3 == 0 and setAmount / 3 == len(unique(allSets)):
            return "Все билеты набора находятся в отчете"
        else:
            return "Нужно проверить открытки"


class SaleListTicketsInArray(SalesDataBase):
    def search_tickets(self, *ticket_type, is_instant=False):
        tickets = []
        bingo_lottery_codes = ["102021", "102041", "102051", "102091"]

        for i in range(2, self.sheet.max_row):
            product_code = self._get_cell_value('J', i)[0:6]
            ticket_number = self._get_cell_value('J', i)
            report_ticket_type = self._get_cell_value('K', i)

            # Список моментальных и тиражных билетов
            if report_ticket_type in ['Бумажный', 'Открытка'] and report_ticket_type in ticket_type :

                #TODO: не нравятся условия, должен быть другой вариант
                if product_code not in bingo_lottery_codes and is_instant == True:
                    tickets.append(ticket_number)
                elif product_code in bingo_lottery_codes and is_instant != True:
                    tickets.append(ticket_number)

            # Список электронных билетов
            elif report_ticket_type in ticket_type:
                tickets.append(ticket_number)

        return tickets


    def digital_tickets(self):
        digitals = self.search_tickets('Электронный', 'Купон')
        return digitals

    def draw_tickets(self):
        draw_tickets = self.search_tickets('Бумажный', 'Открытка')
        return draw_tickets

    def instant_tickets(self):
        instants = self.search_tickets('Бумажный', is_instant=True)
        return instants
