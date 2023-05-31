from openpyxl import load_workbook
import files_path
from openpyxl.utils import get_column_letter
from payment_report import PaymentsAsserts
from util import _get_cell_value, get_boundary_values


class CommonFunctions:
    def __init__(self):
        # Основная таблица
        self.file_path = files_path.payment_registry
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active
        # Проверка наличия доп. листа в файле
        if "Реестр выплаченных выигрышей" not in self.book.sheetnames:
            self.book.create_sheet("Реестр выплаченных выигрышей")
        self.add_sheet = self.book["Реестр выплаченных выигрышей"]

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)


class PreCondition(CommonFunctions):
    """
    Класс для выполнения предусловия по копированию части таблицы
    и дальнейшее ее редактирование
    """
    def copying_table(self):
        self.add_sheet.delete_rows(1, self.add_sheet.max_row)
        # Копирование основной части файла в итоговый отчет
        begin_column = int(self.sheet.max_column) - 3
        second_last_page = self.sheet.iter_rows(min_col=begin_column - 4, max_col=begin_column - 1, values_only=True)
        last_page = self.sheet.iter_rows(min_col=begin_column, values_only=True)

        # Копирование строк в новый файл
        for row in second_last_page:
            self.add_sheet.append(row)
        for row in last_page:
            self.add_sheet.append(row)

        self.book.save(self.file_path)

    def delete_rows(self):
        # Удаление лишних строк в начале
        first_row = 0
        for row in list(self.add_sheet.rows):
            for cell in row:
                if cell.value == "Итого по каждой игре":
                    first_row = cell.row
                    break
            if first_row > 0:
                break

        if first_row > 0:
            self.add_sheet.delete_rows(1, first_row)

        # Удаление лишних строк в конце
        last_row = 0
        for row in reversed(list(self.add_sheet.rows)):
            for cell in row:
                if cell.value == "ИТОГО:":
                    last_row = cell.row
                    break
            if last_row > 0:
                break

        if last_row > 0:
            self.add_sheet.delete_rows(last_row + 1, self.add_sheet.max_row - last_row)

        self.book.save(self.file_path)


class ReportPaidPaymentRegistry(CommonFunctions):
    def paid_lottery_names_list(self):
        return list(self.add_sheet.rows)

    def finalInfo(self, value):

        info = {
            "length": int(len(self.paid_lottery_names_list())),
            "number": float(self.add_sheet["C" + str(self.add_sheet.max_row)].value),
            "amount": float(self.add_sheet["D" + str(self.add_sheet.max_row)].value.replace(" ", ""))
        }

        return info[value]

    #TODO: добавить в отчет для проверки

    # Поиск расхождение между реестром и отчетом по выплатам
    def discrepancy_reports(self):
        global ticket_number_cell
        remaining_tickets = PaymentsAsserts.collecting_numbers()
        try:
            for number in range(3, int(self.sheet.max_column), 4):
                ticket_numbers_column = get_column_letter(number)
                win_amounts_column = get_column_letter(number + 1)

                for row in range(1, self.sheet.max_row + 1):
                    ticket_number_cell = self.sheet[ticket_numbers_column + str(row)].value
                    win_amount_cell = self.sheet[win_amounts_column + str(row)].value

                    if ticket_number_cell is not None and type(
                            ticket_number_cell) != float and ticket_number_cell.isdigit():
                        print(remaining_tickets[ticket_number_cell])
                        if remaining_tickets[ticket_number_cell] == int(win_amount_cell.replace(' ', '')):
                            remaining_tickets.pop(ticket_number_cell)
        except ValueError:
            remaining_tickets = f"Ошибка! Что-то не так с билетом \"{ticket_number_cell}\" " \
                                f"на строке {row}, возможно, нет в отчете выплат!"

        except KeyError:
            remaining_tickets = f"Ошибка! Билета \"{ticket_number_cell}\" " \
                                f"на строке {row} нет в реестре выплаченных выигрышей, но есть в отчете по продажам!"

        return remaining_tickets

    # TODO: добавить в отчет для проверки
    # Счет билетов по столбцу "Номер лотерейного билета..."
    def count_tickets_and_winnings(self):
        amount_tickets = 0
        amount_winnings = 0

        for number in range(3, int(self.sheet.max_column), 4):
            ticket_numbers_column = get_column_letter(number)
            win_amounts_column = get_column_letter(number + 1)

            for row in range(1, self.sheet.max_row + 1):
                ticket_number_cell = self.sheet[ticket_numbers_column + str(row)].value
                win_amount_column = self.sheet[win_amounts_column + str(row)].value

                if ticket_number_cell is not None and type(ticket_number_cell) != int and ticket_number_cell.isdigit():
                    amount_tickets += 1
                    amount_winnings += int(win_amount_column.replace(' ', ''))

        return amount_tickets, amount_winnings
