from openpyxl import load_workbook
from util import _get_cell_value, get_boundary_values
import files_path

class CommonFunctions:
    def __init__(self):
        self.file_path = files_path.branch_report
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)


class WorkSheet(CommonFunctions):
    def table_range(self, table):
        """Возвращает диапазон выбранной таблицы

        :param table: 'tickets' или 'receipts'
        :return: string
        """
        # задаем начальное и конечное значение для поиска диапазона
        start_value_row = 'Наименование лотереи'
        end_value_row = 'ИТОГО:'
        end_value_column = 'Подлежит перечислению Принципалу за отчетный период'
        table_range = [start_value_row, end_value_column, end_value_row]

        # Функция iter_rows() возвращает все строки из файла
        rows = self.sheet.iter_rows()
        # Переменная `cells` содержит генератор, который является последовательностью всех ячеек в листе таблицы
        cells = (cell for row in rows for cell in row)
        # Поиск стартовой точки, последней колонки и строки
        ranges_in_array = (cell for cell in cells if cell.value in table_range)

        # Диапазон первой таблицы
        start_position = next(ranges_in_array)
        last_column, last_row = next(ranges_in_array), next(ranges_in_array)
        # Диапазон второй таблицы
        start_position_st = next(ranges_in_array)
        last_column_st, last_row_st = next(ranges_in_array), next(ranges_in_array)

        coordinate = {
            'start_row': start_position.row + 4 if table == 'tickets' else start_position_st.row + 4,
            'start_column': start_position.column if table == 'tickets' else start_position_st.column,
            'last_column': last_column.column if table == 'tickets' else last_column_st.column + 4,
            'last_row': last_row.row if table == 'tickets' else last_row_st.row
        }

        start_coordinate = self.sheet.cell(row=coordinate["start_row"],
                                           column=coordinate["start_column"]).coordinate
        end_coordinate = self.sheet.cell(row=coordinate["last_row"], column=coordinate["last_column"]).coordinate
        table_range = start_coordinate + ':' + end_coordinate

        return table_range

# Объекты класса WorkSheet
workSheet = WorkSheet()
tickets_table_range = workSheet.table_range('tickets')
receipts_table_range = workSheet.table_range('receipts')

class ReportBranchData(CommonFunctions):
    # Получение данных из "ИТОГО" в таблице "Реализация лотерейных билетов"
    def total_values_lottery_tickets(self, column):
        result_row = get_boundary_values(tickets_table_range, 'max_row')
        values = {
            'sold_number': self._get_cell_value('H', result_row),
            'sold_amount': self._get_cell_value('I', result_row),
            'paid_number': self._get_cell_value('L', result_row),
            'paid_amount': self._get_cell_value('M', result_row),
            'reward': round(float(self._get_cell_value('P', result_row)), 2),
            'transfer': round(float(self._get_cell_value('Q', result_row)), 2)
        }
        return values.get(column)

    # Получение данных из "ИТОГО" в таблице "Реализация лотерейных квитанций"
    def total_values_lottery_receipts(self, column):
        result_row = get_boundary_values(receipts_table_range, 'max_row')
        values = {
            'sold_number': self._get_cell_value('C', result_row),
            'sold_amount': self._get_cell_value('E', result_row),
            'paid_number': self._get_cell_value('H', result_row),
            'paid_amount': self._get_cell_value('J', result_row),
            'reward': round(float(self._get_cell_value('N', result_row)), 2),
            'transfer': round(float(self._get_cell_value('P', result_row)), 2)
        }
        return values.get(column)

    # Общая сумма двух таблиц по вознаграждению
    def reward_of_two_tables(self):
        result_row = int(get_boundary_values(receipts_table_range, 'max_row')) + 3
        reward = self._get_cell_value('I', result_row)
        return reward

    # Общая сумма двух таблиц по перечислению средств
    def transfer_of_two_tables(self):
        result_row = int(get_boundary_values(receipts_table_range, 'max_row')) + 4
        reward = self._get_cell_value('I', result_row)
        return reward


class BranchAsserts(CommonFunctions):
    """
        Функция для подсчета итоговых данных
        в таблицах "Реализация лотерейных билетов" и "Реализация лотерейных квитанций"
    """
    def counting_values_in_column(self, table, column, data_type='int'):
        total = 0
        if table == 'realization_tickets':
            diapason = tickets_table_range
        elif table == 'realization_receipts':
            diapason = receipts_table_range
        else:
            raise ValueError('Неверный тип таблицы')

        min_row = int(get_boundary_values(diapason, 'min_row'))
        max_row = int(get_boundary_values(diapason, 'max_row'))

        for row in range(min_row, max_row):
            cell = self.sheet.cell(row=row, column=column).value
            if data_type == 'float':
                total += float(cell)
            else:
                total += int(cell)

        if data_type == 'float':
            return round(total, 2)
        else:
            return total

    # "Реализация лотерейных билетов"
    def sold_number_tickets(self):
        return self.counting_values_in_column('realization_tickets', 8)

    def sold_amount_tickets(self):
        return self.counting_values_in_column('realization_tickets', 9)

    def paid_number_tickets(self):
        return self.counting_values_in_column('realization_tickets', 12)

    def paid_amount_tickets(self):
        return self.counting_values_in_column('realization_tickets', 13)

    def reward_tickets(self):
        return self.counting_values_in_column('realization_tickets', 16, 'float')

    def transfer_tickets(self):
        return self.counting_values_in_column('realization_tickets', 17, 'float')

    # "Реализация лотерейных квитанций"
    def sold_number_receipts(self):
        return self.counting_values_in_column('realization_receipts', 3)

    def sold_amount_receipts(self):
        return self.counting_values_in_column('realization_receipts', 5)

    def paid_number_receipts(self):
        return self.counting_values_in_column('realization_receipts', 8)

    def paid_amount_receipts(self):
        return self.counting_values_in_column('realization_receipts', 10)

    def reward_receipts(self):
        return self.counting_values_in_column('realization_receipts', 14, 'float')

    def transfer_receipts(self):
        return self.counting_values_in_column('realization_receipts', 16, 'float')

    # Общее вознаграждение и перечисление
    def total_rewards(self):
        tickets = self.reward_tickets()
        receipts = self.reward_receipts()
        return tickets + receipts

    def total_transfer(self):
        tickets = self.transfer_tickets()
        receipts = self.transfer_receipts()
        return tickets + receipts

    # Проверка расчетов в каждой строке
    def check_row(self, table):
        if table == 'realization_tickets':
            diapason = tickets_table_range
        elif table == 'realization_receipts':
            diapason = receipts_table_range
        else:
            raise ValueError('Неверный тип таблицы')

        min_row = int(get_boundary_values(diapason, 'min_row'))
        max_row = int(get_boundary_values(diapason, 'max_row'))

        for row in range(min_row, max_row):
            columns = {
                'ticket_price_column': 4 if table == 'realization_tickets' else 1,
                'sold_number_column': 8 if table == 'realization_tickets' else 3,
                'sold_amount_column': 9 if table == 'realization_tickets' else 5,
                'paid_amount_column': 13 if table == 'realization_tickets' else 10,
                'percent_column': 15 if table == 'realization_tickets' else 13,
                'reward_column': 16 if table == 'realization_tickets' else 14,
                'transfer_column': 17 if table == 'realization_tickets' else 16,
            }

            ticket_price = self.sheet.cell(row=row, column=columns['ticket_price_column']).value
            sold_number = self.sheet.cell(row=row, column=columns['sold_number_column']).value
            sold_amount = self.sheet.cell(row=row, column=columns['sold_amount_column']).value
            paid_amount = self.sheet.cell(row=row, column=columns['paid_amount_column']).value
            percent = self.sheet.cell(row=row, column=columns['percent_column']).value
            reward = float(self.sheet.cell(row=row, column=columns['reward_column']).value)
            transfer = float(self.sheet.cell(row=row, column=columns['transfer_column']).value)

            try:
                # Подсчет и сверка
                if table == 'realization_tickets':
                    count_sold_amount = ticket_price * sold_number
                    assert count_sold_amount == sold_amount

                count_reward = round(sold_amount * (percent / 100), 1)
                assert count_reward == round(reward, 1)

                count_transfer = sold_amount - paid_amount - round(reward, 1)
                assert count_transfer == transfer

                return "ДА"
            except:
                return 'НЕТ!\nСтрока ' + str(row)

