from openpyxl import load_workbook
from util import _get_cell_value, get_boundary_values
import files_path


class CommonFunctions:
    def __init__(self):
        self.file_path = files_path.agent_noncirculated_report
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)


class WorkSheet(CommonFunctions):
       def table_range(self):
        # задаем начальное и конечное значение для поиска диапазона
        start_value_row = 'Наименование лотереи'
        end_value_row = 'ИТОГО:'
        end_value_column = 'Остаток на конец  отчетного периода'
        table_range = [start_value_row, end_value_column, end_value_row]

        # Функция iter_rows() возвращает все строки из файла
        rows = self.sheet.iter_rows()
        # Переменная `cells` содержит генератор, который является последовательностью всех ячеек в листе таблицы
        cells = (cell for row in rows for cell in row)
        # Поиск стартовой точки, последней колонки и строки
        ranges_in_array = (cell for cell in cells if cell.value in table_range)

        # Диапазон первой таблицы
        start_position = next(ranges_in_array)
        last_column, last_row = next(ranges_in_array), next(ranges_in_array)

        coordinate = {
            'start_row': start_position.row + 4,
            'start_column': start_position.column,
            'last_column': last_column.column,
            'last_row': last_row.row
        }

        start_coordinate = self.sheet.cell(row=coordinate["start_row"], column=coordinate["start_column"]).coordinate
        end_coordinate = self.sheet.cell(row=coordinate["last_row"], column=coordinate["last_column"]).coordinate
        table_range = start_coordinate + ':' + end_coordinate

        return table_range


# Объекты класса WorkSheet
workSheet = WorkSheet()
tickets_table_range = workSheet.table_range()

class ReportAgentNoncirculatedData(CommonFunctions):
    # Беру данные из "ИТОГО" в таблице "Реализация лотерейных билетов"
    def total_values_lottery_tickets(self, column):
        result_row = get_boundary_values(tickets_table_range, 'max_row')
        # Продажи
        sold_number = self._get_cell_value('H', result_row)
        sold_amount = self._get_cell_value('I', result_row)
        # Выплаты
        paid_number = self._get_cell_value('M', result_row)
        paid_amount = self._get_cell_value('N', result_row)
        # Вознаграждения
        reward = self._get_cell_value('P', result_row)
        # Перечисление принципалу
        transfer = self._get_cell_value('Q', result_row)

        if column == 'sold_number':
            return sold_number
        elif column == 'sold_amount':
            return sold_amount
        elif column == 'paid_number':
            return paid_number
        elif column == 'paid_amount':
            return paid_amount
        elif column == 'reward':
            return reward
        elif column == 'transfer':
            return transfer

    # Итоговые суммы ниже таблицы
    def reward_under_table(self):
        result_row = int(get_boundary_values(tickets_table_range, 'max_row')) + 8
        rubles = self._get_cell_value('G', result_row)
        pennies = self._get_cell_value('J', result_row)/100
        reward = rubles + pennies
        return reward

    def transfer_under_table_to_principal(self):
        result_row = int(get_boundary_values(tickets_table_range, 'max_row')) + 11
        rubles = self._get_cell_value('G', result_row)
        pennies = self._get_cell_value('J', result_row)/100
        reward = rubles + pennies
        return reward

    def transfer_under_table_to_agent(self):
        result_row = int(get_boundary_values(tickets_table_range, 'max_row')) + 14
        rubles = self._get_cell_value('G', result_row)
        pennies = self._get_cell_value('J', result_row)/100
        reward = rubles + pennies
        return reward


class AgentNoncirculatedAsserts(CommonFunctions):
    # Функция для подсчета итоговых данных
    # в таблице "Реализация лотерейных билетов"
    def counting_values_in_column(self, column, data_type='int'):
        min_row = int(get_boundary_values(tickets_table_range, 'min_row'))
        max_row = int(get_boundary_values(tickets_table_range, 'max_row'))

        values = [self.sheet.cell(row=row, column=column).value for row in range(min_row, max_row)]
        total = sum(float(v) if data_type == 'float' else int(v) for v in values)

        return round(total, 2) if data_type == 'float' else total

    # "Реализация лотерейных билетов"
    def sold_number_tickets(self):
        return self.counting_values_in_column(8)

    def sold_amount_tickets(self):
        return self.counting_values_in_column(9)

    def paid_number_tickets(self):
        return self.counting_values_in_column(13)

    def paid_amount_tickets(self):
        return self.counting_values_in_column(14)

    def reward_tickets(self):
        return self.counting_values_in_column(16, 'float')

    def transfer_tickets(self):
        return self.counting_values_in_column(17, 'float')

    # Проверка расчетов в каждой строке
    def check_row(self):
        min_row = int(get_boundary_values(tickets_table_range, 'min_row'))
        max_row = int(get_boundary_values(tickets_table_range, 'max_row'))

        for row in range(min_row, max_row):
            columns = {
                'sold_amount_column': 9,
                'paid_amount_column': 14,
                'percent_column': 15,
                'reward_column': 16,
                'transfer_column': 17
            }

            sold_amount = self.sheet.cell(row=row, column=columns['sold_amount_column']).value
            paid_amount = self.sheet.cell(row=row, column=columns['paid_amount_column']).value
            percent = self.sheet.cell(row=row, column=columns['percent_column']).value
            reward = float(self.sheet.cell(row=row, column=columns['reward_column']).value)
            transfer = float(self.sheet.cell(row=row, column=columns['transfer_column']).value)

            try:
                count_reward = round(sold_amount * (percent / 100), 1)
                assert count_reward == round(reward, 1)

                count_transfer = sold_amount - paid_amount - round(reward, 1)
                assert count_transfer == transfer

                return "ДА"
            except:
                return 'НЕТ!\nСтрока ' + str(row)
