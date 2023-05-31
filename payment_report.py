from openpyxl import load_workbook
from numpy import unique
from util import _get_cell_value
import files_path


class CommonFunctions:

    def __init__(self):
        self.file_path = files_path.payment_report
        self.book = load_workbook(filename=self.file_path, data_only=True)
        self.sheet = self.book.active

    def _get_cell_value(self, column_letter, idx):
        return _get_cell_value(column_letter, idx, self.sheet)


class ReportPaymentsData(CommonFunctions):

    def total_quantity_tickets_in_report(self):
        return self.sheet['M' + str(self.sheet.max_row)].value

    def win_amount_in_report(self):
        return self.sheet['N' + str(self.sheet.max_row)].value

    def check_cells_in_row(self):
        passedAssertions = 0
        bingoLottery = ["102021", "102041", "102051", "102091"]

        for i in range(2, self.sheet.max_row):
            setCell = self.sheet['K' + str(i)].value
            drawCell = self.sheet['H' + str(i)].value
            ticketType = self.sheet['J' + str(i)].value
            seventhDigit = self.sheet['I' + str(i)].value[6:7]
            productCode = self.sheet['I' + str(i)].value[0:6]
            paymentType = self.sheet['L' + str(i)].value

            try:
                if productCode in bingoLottery and seventhDigit == "3" and len(drawCell) == 6 and setCell is not None:
                    assert ticketType == "Открытка" and paymentType == "По билету"
                    passedAssertions += 1

                elif productCode in bingoLottery and seventhDigit == "3" and len(drawCell) == 6 and setCell is None:
                    assert ticketType == "Бумажный" and paymentType == "По билету"
                    passedAssertions += 1

                elif productCode not in bingoLottery and setCell is None and drawCell is None:
                    assert ticketType == "Бумажный" and paymentType == "По билету"
                    passedAssertions += 1

                elif seventhDigit in ["1", "2"] and len(drawCell) == 6 and setCell is None:
                    assert ticketType in ["Электронный", "Купон"]
                    assert paymentType in ["По билету", "По билету с проверочными кодом", "По номеру телефона"]
                    passedAssertions += 1

                else:
                    print(
                        f'Ошибка! Отчет выплат. Строка - {str(i)}\n{self.sheet["I" + str(i)].value}-{len(drawCell)}{setCell}')

            except AssertionError:
                print('Ошибка! Отчет выплат. Строка - ' + str(i))

        return passedAssertions


class PaymentsAsserts(CommonFunctions):

    def total_quantity(self):
        tickets_amount = 0
        for i in range(2, self.sheet.max_row):
            tickets_amount += self.sheet['M' + str(i)].value
        return tickets_amount

    def win_amount(self):
        win_amount = 0
        for i in range(2, self.sheet.max_row):
            win_amount += self.sheet['N' + str(i)].value
        return win_amount

    def counting_tickets(self, type_of_tickets, type_of_lottery=True):
        ticketsQuantity = 0
        winAmount = 0
        bingoLottery = ["102021", "102041", "102051", "102091"]

        for i in range(2, self.sheet.max_row):
            productCode = self.sheet['I' + str(i)].value[0:6]
            ticketType = self.sheet['J' + str(i)].value
            winAmountCell = self.sheet['N' + str(i)].value

            # Подсчет электронных билетов и купонов
            if type_of_tickets == 'Электронный':
                if ticketType in ['Электронный', 'Купон']:
                    ticketsQuantity += 1
                    winAmount += winAmountCell

            # Подсчет бумажных моменталок
            elif type_of_tickets == 'Бумажный' and type_of_lottery == 'Моментальная':
                if ticketType == type_of_tickets and productCode not in bingoLottery:
                    ticketsQuantity += 1
                    winAmount += winAmountCell

            # Подсчет тиражных билетов
            elif type_of_tickets == 'Бумажный' and type_of_lottery == 'Тиражная':
                if ticketType in ['Бумажный', 'Открытка'] and productCode in bingoLottery:
                    ticketsQuantity += 1
                    winAmount += winAmountCell

            # Подсчет бумажных
            elif type_of_tickets == 'Бумажный':
                if ticketType == type_of_tickets:
                    ticketsQuantity += 1
                    winAmount += winAmountCell

            # Подсчет отстальных (т.е. Открыток)
            else:
                if ticketType == type_of_tickets:
                    ticketsQuantity += 1
                    winAmount += winAmountCell

        return ticketsQuantity, winAmount

    def all_ticket_types(self):
        ticketsQuantity = self.counting_tickets('Электронный')[0] + self.counting_tickets('Бумажный')[0] + \
                          self.counting_tickets('Открытка')[0]
        winAmounts = self.counting_tickets('Электронный')[1] + self.counting_tickets('Бумажный')[1] + \
            self.counting_tickets('Открытка')[1]

        return ticketsQuantity, winAmounts

    def unique_ticket_numbers(self):
        ticketNumbersQuantity = 0
        ticketNumbers = []
        for i in range(2, self.sheet.max_row):
            ticketNumber = self.sheet['I' + str(i)].value
            if ticketNumber is not None:
                ticketNumbers.append(ticketNumber)
                ticketNumbersQuantity += 1

        if ticketNumbersQuantity == len(unique(ticketNumbers)):
            return "Все билеты в отчете уникальные"
        else:
            return "В отчете есть дубликат билета"

    def win_amount_less_15000(self):
        winAmountMore15000 = 0

        for i in range(2, self.sheet.max_row):
            ticketSetNumber = self.sheet['N' + str(i)].value
            if ticketSetNumber >= 15000:
                winAmountMore15000 += 1

        if winAmountMore15000 == 0:
            return "Отсутствуют билеты с выигрышем более 15000 руб"
        else:
            return "В отчете есть билеты с выигрышем более 15000 руб"

    # Необходим для paid_payments_registry
    def collecting_numbers(self):
        arrayNumbers = {}

        for i in range(2, self.sheet.max_row):
            ticketNumber = self.sheet['I' + str(i)].value
            winAmount = int(self.sheet['N' + str(i)].value)

            arrayNumbers[ticketNumber] = winAmount

        return arrayNumbers


class PaymentListTicketsInArray(CommonFunctions):
    def search_tickets(self, *ticket_type, is_instant=False):
        tickets = []
        bingo_lottery_codes = ["102021", "102041", "102051", "102091"]

        for i in range(2, self.sheet.max_row):
            product_code = self.sheet['I' + str(i)].value[0:6]
            report_ticket_type = self.sheet['J' + str(i)].value
            ticket_number = self._get_cell_value('I', i)

            # Список моментальных и тиражных билетов
            if report_ticket_type in ['Бумажный', 'Открытка'] and report_ticket_type in ticket_type:

                #TODO: не нравятся условия, должен быть другой вариант
                if product_code not in bingo_lottery_codes and is_instant == True:
                    tickets.append(ticket_number)
                elif product_code in bingo_lottery_codes and is_instant != True:
                    tickets.append(ticket_number)

            # Список электронных билетов
            elif report_ticket_type in ticket_type:
                tickets.append(ticket_number)

        return tickets

    def digital_tickets(self):
        digitals = self.search_tickets('Электронный', 'Купон')
        return digitals

    def draw_tickets(self):
        draw_tickets = self.search_tickets('Бумажный', 'Открытка')
        return draw_tickets

    def instant_tickets(self):
        instants = self.search_tickets('Бумажный', is_instant=True)
        return instants
